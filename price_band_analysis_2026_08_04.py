#!/usr/bin/env python3
"""Generate a log-log price band analysis chart from customer pricing data.

The script reads a CSV or Excel file containing customer name, procedure
quantity, and procedure price columns.  It standardizes the input columns,
fits a trend line in log-log space, overlays one- and two-standard-deviation
bands, displays the scatter plot on logarithmic axes, and exports the full
analysis table to a formatted Excel file.

Usage:
    python price_band_analysis.py
    python price_band_analysis.py input_file.csv
    python price_band_analysis.py input_file.xlsx --sheet "Sheet1"

Dependencies:
    pandas, numpy, matplotlib, openpyxl (for .xlsx input/output)
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

try:
    import tkinter as tk
    from tkinter import filedialog, messagebox
except ImportError:
    tk = None
    filedialog = None
    messagebox = None


# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

# Accepted aliases for each required input column (order = match priority).
COLUMN_ALIASES: dict[str, list[str]] = {
    "customer_name": [
        "customer name",
        "customer group",
        "customer group l3",
        "business partner",
        "customer",
        "account",
        "account name",
        "customer_name",
        "name",
    ],
    "procedure_quantity": [
        "procedure quantity",
        "procedure units",
        "quantity",
        "procedure qty",
        "qty",
        "volume",
        "procedure_quantity",
    ],
    "procedure_price": [
        "procedure price",
        "price",
        "unit price",
        "avg price",
        "procedure_price",
        "net procedure price"
    ],
}

# Display names used after column standardization.
STANDARD_COLUMN_NAMES: dict[str, str] = {
    "customer_name": "Customer Name",
    "procedure_quantity": "Procedure Quantity",
    "procedure_price": "Procedure Price",
}

# Approximate plot-area dimensions (in typographic points) for a (12, 8)-inch
# figure.  Used solely to convert pixel offsets into proportional log-space
# distances for label-collision scoring.  Exact values are not critical; they
# preserve the aspect ratio so that x and y penalties are balanced.
_PLOT_WIDTH_PTS: float = 650.0
_PLOT_HEIGHT_PTS: float = 450.0

CONFIG: dict = {
    # Data-filtering thresholds
    "min_procedure_units": 10,
    "max_procedure_price": 8000,
    # Bin configuration for within-bin outlier analysis
    "max_procedure_unit_bins": 4,
    "min_records_per_bin": 5,
    # When True, Pricing Position is left blank for points within 1 SD
    "band_blank_within_1sd_pricing_position": True,
    # Points beyond this many SDs from the trend line receive scatter labels
    "scatter_label_sd_threshold": 1.0,
    # Within-bin outlier-detection parameters
    "robust_zscore_threshold": 3.5,
    "robust_iqr_multiplier": 1.5,
    "percentile_lower": 0.05,
    "percentile_upper": 0.95,
    # Excel output settings
    "default_output_filename": "price_band_outliers.xlsx",
    "worksheet_name": "Price Band Analysis",
    "header_fill_color": "002060",
    "header_font_color": "FFFFFF",
    "severity_fill_colors": {
        "Normal": "DDEBF7",
        "Lower Concern Outlier": "FCE4D6",
        "High Concern Outlier": "F8CBAD",
    },
    # Percentile-based outlier flag thresholds (0-100 scale).
    # Customers at or below the low threshold on BOTH procedure units and
    # procedure price are flagged as "Outlier - Low Price and Volume".
    # Customers at or above the high threshold on both are flagged as
    # "Outlier - High Price and Volume".
    "outlier_low_percentile_threshold": 30,
    "outlier_high_percentile_threshold": 70,
}


# ---------------------------------------------------------------------------
# Column-name helpers
# ---------------------------------------------------------------------------

def normalize_column_name(column_name: str) -> str:
    """Collapse whitespace, strip, and lowercase a column name for matching."""
    return " ".join(str(column_name).strip().lower().replace("_", " ").split())


def find_matching_column(columns: pd.Index, aliases: list[str]) -> str:
    """Return the first column whose normalized name matches an alias.

    Raises ``KeyError`` if no alias matches any column in *columns*.
    """
    normalized_map = {
        normalize_column_name(col): col for col in columns
    }
    for alias in aliases:
        normalized_alias = normalize_column_name(alias)
        if normalized_alias in normalized_map:
            return normalized_map[normalized_alias]

    raise KeyError(f"Could not find a matching column for aliases: {aliases}")


# ---------------------------------------------------------------------------
# GUI message helpers
# ---------------------------------------------------------------------------

def _show_tk_message(title: str, message: str, *, level: str = "info") -> None:
    """Display a message via a tkinter dialog, falling back to the console.

    Parameters
    ----------
    title : str
        Dialog / console prefix.
    message : str
        Body text.
    level : str
        ``"info"`` for informational messages (stdout) or ``"warning"`` for
        warnings (stderr).
    """
    if messagebox and tk:
        root = tk.Tk()
        root.withdraw()
        try:
            dialog_fn = (
                messagebox.showwarning if level == "warning"
                else messagebox.showinfo
            )
            dialog_fn(title, message)
        finally:
            root.destroy()
    else:
        stream = sys.stderr if level == "warning" else sys.stdout
        print(f"{title}: {message}", file=stream)


def show_info_message(title: str, message: str) -> None:
    """Display an informational message (GUI dialog or console)."""
    _show_tk_message(title, message, level="info")


def show_warning_message(title: str, message: str) -> None:
    """Display a warning message (GUI dialog or console)."""
    _show_tk_message(title, message, level="warning")


# ---------------------------------------------------------------------------
# File-dialog helpers
# ---------------------------------------------------------------------------

def choose_input_file() -> Path:
    """Open a file-selection dialog and return the chosen input path."""
    if not tk or not filedialog:
        raise RuntimeError(
            "File dialog is not available in this Python environment. "
            "Please provide the input file as a command-line argument."
        )

    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)
    try:
        selected_file = filedialog.askopenfilename(
            title="Select input file for price band analysis",
            filetypes=[
                ("Data files", "*.csv *.xlsx *.xls"),
                ("CSV files", "*.csv"),
                ("Excel files", "*.xlsx *.xls"),
                ("All files", "*.*"),
            ],
        )
    finally:
        root.destroy()

    if not selected_file:
        raise ValueError("No input file selected.")

    return Path(selected_file)


def choose_output_file(default_name: str) -> Path:
    """Open a save-as dialog and return the chosen Excel output path."""
    if not tk or not filedialog:
        raise RuntimeError(
            "Save dialog is not available in this Python environment. "
            "Please provide the output file as a command-line argument."
        )

    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)
    try:
        selected_file = filedialog.asksaveasfilename(
            title="Select location for Excel output",
            defaultextension=".xlsx",
            initialfile=default_name,
            filetypes=[
                ("Excel files", "*.xlsx"),
                ("All files", "*.*"),
            ],
        )
    finally:
        root.destroy()

    if not selected_file:
        raise ValueError("No output file selected.")

    return Path(selected_file)


# ---------------------------------------------------------------------------
# Data loading
# ---------------------------------------------------------------------------

def load_excel_sheet(
    file_path: Path,
    sheet_name: str | None = None,
) -> pd.DataFrame:
    """Load a specific Excel sheet or auto-detect the most likely data sheet.

    When *sheet_name* is ``None`` the function checks each sheet for a column
    that matches one of the ``procedure_quantity`` aliases.  The first match
    wins; if none match the first sheet is returned.
    """
    if sheet_name:
        return pd.read_excel(file_path, sheet_name=sheet_name)

    workbook = pd.read_excel(file_path, sheet_name=None)
    quantity_aliases = {
        normalize_column_name(a)
        for a in COLUMN_ALIASES["procedure_quantity"]
    }

    for candidate_df in workbook.values():
        sheet_columns = {
            normalize_column_name(c) for c in candidate_df.columns
        }
        if quantity_aliases & sheet_columns:
            return candidate_df

    # No sheet explicitly matched -- fall back to the first sheet.
    return workbook[next(iter(workbook))]


def load_input_file(
    file_path: Path,
    sheet_name: str | None = None,
) -> pd.DataFrame:
    """Load a CSV or Excel input file into a DataFrame."""
    suffix = file_path.suffix.lower()
    if suffix == ".csv":
        return pd.read_csv(file_path)
    if suffix in {".xlsx", ".xls"}:
        return load_excel_sheet(file_path, sheet_name)

    raise ValueError("Unsupported file type.  Use CSV or Excel (.xlsx/.xls).")


# ---------------------------------------------------------------------------
# Data cleaning & standardization
# ---------------------------------------------------------------------------

def standardize_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Map source column names to the standard analysis column names."""
    mapped = {
        key: find_matching_column(df.columns, aliases)
        for key, aliases in COLUMN_ALIASES.items()
    }

    out = df[
        [mapped["customer_name"], mapped["procedure_quantity"], mapped["procedure_price"]]
    ].copy()
    out.columns = [
        STANDARD_COLUMN_NAMES["customer_name"],
        STANDARD_COLUMN_NAMES["procedure_quantity"],
        STANDARD_COLUMN_NAMES["procedure_price"],
    ]
    return out


def clean_input_data(df: pd.DataFrame) -> pd.DataFrame:
    """Standardize, validate, and filter the raw input for analysis.

    Steps:
      1. Map source columns to standard names.
      2. Coerce quantity and price to numeric (non-numeric becomes NaN).
      3. Drop rows with NaN quantity/price, quantities below the configured
         minimum, non-positive prices, and prices above the configured max.
      4. Notify the user of any dropped rows.
    """
    cleaned = standardize_columns(df)
    customer_col = STANDARD_COLUMN_NAMES["customer_name"]
    quantity_col = STANDARD_COLUMN_NAMES["procedure_quantity"]
    price_col = STANDARD_COLUMN_NAMES["procedure_price"]

    # Normalize customer names; replace blanks with a placeholder.
    cleaned[customer_col] = cleaned[customer_col].astype(str).str.strip()
    cleaned.loc[
        cleaned[customer_col].isin(["", "nan", "None"]),
        customer_col,
    ] = "Unknown"

    # Coerce to numeric and count rows that had non-numeric values.
    cleaned[quantity_col] = pd.to_numeric(cleaned[quantity_col], errors="coerce")
    cleaned[price_col] = pd.to_numeric(cleaned[price_col], errors="coerce")

    non_numeric_mask = cleaned[quantity_col].isna() | cleaned[price_col].isna()
    non_numeric_count = int(non_numeric_mask.sum())

    # Apply all row-level filters in a single pass to avoid repeated copies.
    # NaN comparisons evaluate to False, so non-numeric rows are excluded
    # implicitly by the range conditions.
    valid_mask = (
        ~non_numeric_mask
        & (cleaned[quantity_col] >= CONFIG["min_procedure_units"])
        & (cleaned[price_col] > 0)
        & (cleaned[price_col] <= CONFIG["max_procedure_price"])
    )
    cleaned = cleaned.loc[valid_mask].copy()

    if cleaned.empty:
        message = (
            "No valid data remains after cleaning and filtering.  "
            f"Records must have Procedure Quantity >= {CONFIG['min_procedure_units']} "
            f"and Procedure Price <= {CONFIG['max_procedure_price']}."
        )
        if non_numeric_count > 0:
            message += f"\n\nRows with non-numeric values: {non_numeric_count}"
        show_warning_message("No Valid Data", message)
        raise ValueError("No valid rows found after cleaning input data.")

    if non_numeric_count > 0:
        show_info_message(
            "Data Cleaning Complete",
            f"Loaded successfully.\nRows dropped (non-numeric values): {non_numeric_count}",
        )

    return cleaned


# ---------------------------------------------------------------------------
# Trend fitting
# ---------------------------------------------------------------------------

def fit_log_log_trend(
    df: pd.DataFrame,
) -> tuple[np.ndarray, np.ndarray, np.ndarray, float]:
    """Fit a linear regression in log10-log10 space.

    Returns
    -------
    x_values : ndarray
        Raw procedure quantities.
    y_values : ndarray
        Raw procedure prices.
    trend_values : ndarray
        Predicted prices on the original (non-log) scale.
    residual_std : float
        Standard deviation of log10 residuals (ddof=1).
    """
    quantity_col = STANDARD_COLUMN_NAMES["procedure_quantity"]
    price_col = STANDARD_COLUMN_NAMES["procedure_price"]

    x = df[quantity_col].to_numpy(dtype=float)
    y = df[price_col].to_numpy(dtype=float)

    if len(df) < 2:
        raise ValueError("At least two rows are required to fit a trend line.")

    log_x, log_y = np.log10(x), np.log10(y)
    slope, intercept = np.polyfit(log_x, log_y, 1)
    trend_log = slope * log_x + intercept
    residual_std = float(np.std(log_y - trend_log, ddof=1)) if len(df) > 2 else 0.0

    return x, y, np.power(10.0, trend_log), residual_std


# ---------------------------------------------------------------------------
# Procedure-unit binning
# ---------------------------------------------------------------------------

def build_procedure_unit_bins(
    quantity_series: pd.Series,
) -> tuple[pd.Series, pd.DataFrame]:
    """Assign each row to a quantile-based procedure-unit bin.

    Returns
    -------
    bin_labels : Series
        Maps each row index to its bin name (``"Bin 1"``, ``"Bin 2"``, ...).
    bin_summary : DataFrame
        Two columns -- *Procedure Unit Bin* and *Procedure Unit Bin Range*
        (human-readable min-to-max string).
    """
    bin_count = min(CONFIG["max_procedure_unit_bins"], max(len(quantity_series), 1))

    if bin_count <= 1:
        bin_labels = pd.Series(
            ["Bin 1"] * len(quantity_series),
            index=quantity_series.index,
        )
    else:
        rank_values = quantity_series.rank(method="first")
        raw_bins = pd.qcut(rank_values, q=bin_count, labels=False, duplicates="drop")
        if raw_bins is None:
            bin_labels = pd.Series(
                ["Bin 1"] * len(quantity_series),
                index=quantity_series.index,
            )
        else:
            bin_labels = raw_bins.astype(int).add(1).map(lambda v: f"Bin {v}")

    bin_summary = (
        pd.DataFrame({
            "Procedure Unit Bin": bin_labels,
            "Procedure Quantity": quantity_series,
        })
        .groupby("Procedure Unit Bin", sort=True)["Procedure Quantity"]
        .agg(["min", "max"])
        .reset_index()
    )
    bin_summary["Procedure Unit Bin Range"] = bin_summary.apply(
        lambda row: f"{row['min']:,.0f} to {row['max']:,.0f}",
        axis=1,
    )

    return bin_labels, bin_summary[["Procedure Unit Bin", "Procedure Unit Bin Range"]]


# ---------------------------------------------------------------------------
# Within-bin outlier detection
# ---------------------------------------------------------------------------

def calculate_robust_z_scores(series: pd.Series) -> pd.Series:
    """Return Modified Z-scores based on the Median Absolute Deviation (MAD).

    The scaling constant 0.6745 is the 75th-percentile of the standard normal
    distribution, making the MAD a consistent estimator of the standard
    deviation under normality.
    """
    median_val = series.median()
    mad = np.median(np.abs(series - median_val))

    if mad == 0 or np.isnan(mad):
        return pd.Series(0.0, index=series.index)

    return 0.6745 * (series - median_val) / mad


def append_bin_outlier_metrics(analysis_df: pd.DataFrame) -> pd.DataFrame:
    """Add procedure-unit bin labels and per-bin outlier diagnostics.

    Three independent methods flag outliers within each bin:

    1. **Robust Z-score** -- modified Z-score exceeding the configured
       threshold (``robust_zscore_threshold``).
    2. **Robust IQR** -- price outside ``Q1 - k*IQR ... Q3 + k*IQR``
       (``robust_iqr_multiplier``).
    3. **Percentile extremes** -- rank below ``percentile_lower`` or above
       ``percentile_upper``.

    Bins with fewer than ``min_records_per_bin`` peers suppress all outlier
    flags to avoid false positives from small samples.
    """
    price_col = STANDARD_COLUMN_NAMES["procedure_price"]
    quantity_col = STANDARD_COLUMN_NAMES["procedure_quantity"]

    enriched = analysis_df.copy()
    bin_labels, bin_summary = build_procedure_unit_bins(enriched[quantity_col])
    enriched["Procedure Unit Bin"] = bin_labels
    enriched = enriched.merge(bin_summary, on="Procedure Unit Bin", how="left")

    # Pre-initialize metric columns with safe defaults so every row has a
    # value even if its bin is skipped or too small for detection.
    enriched["Bin Peer Count"] = 0
    enriched["Bin Median Price"] = np.nan
    enriched["Bin Robust Z Score"] = 0.0
    enriched["Bin Robust Z Score Outlier"] = False
    enriched["Bin IQR Lower Bound"] = np.nan
    enriched["Bin IQR Upper Bound"] = np.nan
    enriched["Bin Robust IQR Outlier"] = False
    enriched["Bin Percentile Rank"] = np.nan
    enriched["Bin Lower Percentile Threshold"] = np.nan
    enriched["Bin Upper Percentile Threshold"] = np.nan
    enriched["Bin Percentile Extreme"] = False
    enriched["Bin Outlier Methods Triggered"] = 0
    enriched["Bin Outlier Flag"] = False
    enriched["Bin Outlier Methods"] = ""

    for _, bin_idx in enriched.groupby("Procedure Unit Bin").groups.items():
        prices = enriched.loc[bin_idx, price_col]
        peer_count = len(prices)

        # -- Descriptive statistics ----------------------------------------
        q1, q3 = prices.quantile(0.25), prices.quantile(0.75)
        iqr = q3 - q1
        lower_bound = q1 - CONFIG["robust_iqr_multiplier"] * iqr
        upper_bound = q3 + CONFIG["robust_iqr_multiplier"] * iqr
        robust_z = calculate_robust_z_scores(prices)
        pct_ranks = prices.rank(method="average", pct=True)
        lower_thresh = prices.quantile(CONFIG["percentile_lower"])
        upper_thresh = prices.quantile(CONFIG["percentile_upper"])

        # -- Outlier masks for each method ---------------------------------
        z_outlier = robust_z.abs() > CONFIG["robust_zscore_threshold"]
        iqr_outlier = (prices < lower_bound) | (prices > upper_bound)
        pct_outlier = (
            (pct_ranks <= CONFIG["percentile_lower"])
            | (pct_ranks >= CONFIG["percentile_upper"])
        )

        # Suppress all flags when the bin is too small for reliable detection.
        if peer_count < CONFIG["min_records_per_bin"]:
            z_outlier = pd.Series(False, index=prices.index)
            iqr_outlier = pd.Series(False, index=prices.index)
            pct_outlier = pd.Series(False, index=prices.index)

        methods_triggered = (
            z_outlier.astype(int)
            + iqr_outlier.astype(int)
            + pct_outlier.astype(int)
        )

        # Build a human-readable comma-separated list of which methods fired.
        method_labels = pd.DataFrame({
            "robust_z": z_outlier,
            "robust_iqr": iqr_outlier,
            "percentile_extreme": pct_outlier,
        }).apply(
            lambda row: ", ".join(
                name
                for name, flag in [
                    ("Robust Z score", row["robust_z"]),
                    ("Robust IQR", row["robust_iqr"]),
                    ("Percentile extremes", row["percentile_extreme"]),
                ]
                if flag
            ),
            axis=1,
        )

        # -- Assign computed metrics back to the main DataFrame ------------
        enriched.loc[bin_idx, "Bin Peer Count"] = peer_count
        enriched.loc[bin_idx, "Bin Median Price"] = prices.median()
        enriched.loc[bin_idx, "Bin Robust Z Score"] = robust_z
        enriched.loc[bin_idx, "Bin Robust Z Score Outlier"] = z_outlier
        enriched.loc[bin_idx, "Bin IQR Lower Bound"] = lower_bound
        enriched.loc[bin_idx, "Bin IQR Upper Bound"] = upper_bound
        enriched.loc[bin_idx, "Bin Robust IQR Outlier"] = iqr_outlier
        enriched.loc[bin_idx, "Bin Percentile Rank"] = pct_ranks
        enriched.loc[bin_idx, "Bin Lower Percentile Threshold"] = lower_thresh
        enriched.loc[bin_idx, "Bin Upper Percentile Threshold"] = upper_thresh
        enriched.loc[bin_idx, "Bin Percentile Extreme"] = pct_outlier
        enriched.loc[bin_idx, "Bin Outlier Methods Triggered"] = methods_triggered
        enriched.loc[bin_idx, "Bin Outlier Flag"] = methods_triggered > 0
        enriched.loc[bin_idx, "Bin Outlier Methods"] = method_labels

    return enriched


def append_percentile_outlier_flags(df: pd.DataFrame) -> pd.DataFrame:
    """Add procedure-unit and price percentile ranks with an outlier flag.

    Percentile ranks are computed across the full dataset using the
    PERCENTRANK.INC convention (0 for the minimum value, 100 for the
    maximum).  Results are rounded to one decimal place.

    The outlier flag marks customers as:

    * ``"Outlier - Low Price and Volume"`` when both the units percentile
      and the price percentile are at or below
      ``CONFIG["outlier_low_percentile_threshold"]``.
    * ``"Outlier - High Price and Volume"`` when both are at or above
      ``CONFIG["outlier_high_percentile_threshold"]``.
    """
    quantity_col = STANDARD_COLUMN_NAMES["procedure_quantity"]
    price_col = STANDARD_COLUMN_NAMES["procedure_price"]

    enriched = df.copy()
    n = len(enriched)

    if n > 1:
        qty_rank = enriched[quantity_col].rank(method="min")
        price_rank = enriched[price_col].rank(method="min")
        enriched["Procedure Units Percentile"] = (
            (qty_rank - 1) / (n - 1) * 100
        ).round(1)
        enriched["Procedure Price Percentile"] = (
            (price_rank - 1) / (n - 1) * 100
        ).round(1)
    else:
        enriched["Procedure Units Percentile"] = 0.0
        enriched["Procedure Price Percentile"] = 0.0

    low_thresh = CONFIG["outlier_low_percentile_threshold"]
    high_thresh = CONFIG["outlier_high_percentile_threshold"]

    low_mask = (
        (enriched["Procedure Units Percentile"] <= low_thresh)
        & (enriched["Procedure Price Percentile"] <= low_thresh)
    )
    high_mask = (
        (enriched["Procedure Units Percentile"] >= high_thresh)
        & (enriched["Procedure Price Percentile"] >= high_thresh)
    )

    enriched["Outlier Flag"] = ""
    enriched.loc[low_mask, "Outlier Flag"] = "Outlier - Low Price and Volume"
    enriched.loc[high_mask, "Outlier Flag"] = "Outlier - High Price and Volume"

    return enriched


# ---------------------------------------------------------------------------
# Outlier classification (trend-based)
# ---------------------------------------------------------------------------

def classify_outliers(
    df: pd.DataFrame,
    trend_values: np.ndarray,
    residual_std: float,
) -> pd.DataFrame:
    """Classify each record by its deviation from the log-log trend.

    Band assignment:
      * **Within 1 SD** -- Normal
      * **Between 1 and 2 SD** -- Lower Concern Outlier
      * **Outside 2 SD** -- High Concern Outlier

    ``Pricing Position`` (``"Potentially Underpriced"`` /
    ``"Potentially Overpriced"``) is assigned only to points outside 1 SD.
    """
    price_col = STANDARD_COLUMN_NAMES["procedure_price"]

    out = df.copy()
    out["Trend Price"] = trend_values
    out["Log Residual"] = np.log10(out[price_col]) - np.log10(out["Trend Price"])
    out["Residual Ratio"] = out[price_col] / out["Trend Price"]
    out["Band"] = "Within 1 SD"
    out["Severity"] = "Normal"
    out["Pricing Position"] = ""

    if residual_std != 0:
        abs_resid = out["Log Residual"].abs()
        between_mask = (abs_resid > residual_std) & (abs_resid <= 2 * residual_std)
        beyond_mask = abs_resid > 2 * residual_std

        out.loc[between_mask, "Band"] = "Between 1 and 2 SD"
        out.loc[beyond_mask, "Band"] = "Outside 2 SD"
        out.loc[between_mask, "Severity"] = "Lower Concern Outlier"
        out.loc[beyond_mask, "Severity"] = "High Concern Outlier"

        outside_1sd = abs_resid > residual_std
        out.loc[
            outside_1sd & (out[price_col] < out["Trend Price"]),
            "Pricing Position",
        ] = "Potentially Underpriced"
        out.loc[
            outside_1sd & (out[price_col] > out["Trend Price"]),
            "Pricing Position",
        ] = "Potentially Overpriced"

    if not CONFIG["band_blank_within_1sd_pricing_position"]:
        out.loc[out["Band"] == "Within 1 SD", "Pricing Position"] = "In Line"

    return append_percentile_outlier_flags(append_bin_outlier_metrics(out))


# ---------------------------------------------------------------------------
# Plotting helpers -- band lines
# ---------------------------------------------------------------------------

def get_band_lines(
    trend_values: np.ndarray,
    residual_std: float,
) -> dict[str, np.ndarray]:
    """Return multiplicative +/-1 SD and +/-2 SD band lines.

    In log-space, adding/subtracting a constant is equivalent to multiplying/
    dividing in linear space, so the bands are symmetric on the log-scale
    chart.
    """
    if residual_std == 0:
        return {}

    f1 = np.power(10.0, residual_std)
    f2 = np.power(10.0, 2 * residual_std)
    return {
        "+1 SD": trend_values * f1,
        "-1 SD": trend_values / f1,
        "+2 SD": trend_values * f2,
        "-2 SD": trend_values / f2,
    }


# ---------------------------------------------------------------------------
# Plotting helpers -- scatter-label placement
# ---------------------------------------------------------------------------

def _generate_candidate_offsets(
    num_angles: int = 12,
    radii: tuple[float, ...] = (14.0, 24.0, 36.0),
) -> list[tuple[float, float]]:
    """Generate candidate label offsets arranged in concentric rings.

    Returns a list of ``(dx, dy)`` pairs in display-point units.  Each ring
    places *num_angles* evenly-spaced directions at the given radius, giving
    ``num_angles * len(radii)`` total candidates (default: 36).  Successive
    rings are offset by half an angular step so that inner and outer
    candidates do not share the same radial line, improving coverage in
    dense clusters.

    Parameters
    ----------
    num_angles : int
        Number of evenly-spaced directions per ring.
    radii : tuple[float, ...]
        Distance of each ring from the anchor, in display points.
    """
    offsets: list[tuple[float, float]] = []
    for ring_idx, radius in enumerate(radii):
        # Stagger alternate rings by half a step for better coverage.
        phase = (ring_idx * np.pi) / num_angles
        for i in range(num_angles):
            angle = phase + 2.0 * np.pi * i / num_angles
            offsets.append((radius * np.cos(angle), radius * np.sin(angle)))
    return offsets


def _select_best_offset(
    anchor_log_x: float,
    anchor_log_y: float,
    x_log_span: float,
    y_log_span: float,
    placed_labels: list[tuple[float, float]],
    data_points_log: np.ndarray,
    candidate_offsets: np.ndarray,
) -> tuple[float, float]:
    """Choose the candidate offset that best avoids visual clutter.

    Scoring (lower is better) combines three terms:

    * **Label repulsion** -- sum of inverse distances to already-placed
      labels.  Heavily weighted to prevent label-on-label overlap.
    * **Point repulsion** -- sum of inverse distances to all scatter-plot
      data points.  Moderately weighted so labels avoid obscuring data.
    * **Anchor distance** -- slight preference for keeping the label close
      to its data point to maintain readability.

    All position comparisons happen in an abstract space proportional to the
    log-scaled axes, so the scoring respects the visual layout of the chart.

    Parameters
    ----------
    anchor_log_x, anchor_log_y : float
        Log10-space coordinates of the data point being labelled.
    x_log_span, y_log_span : float
        Full log10-space range of each axis (used for coordinate conversion).
    placed_labels : list[tuple[float, float]]
        Log-space positions of labels already committed to the chart.
    data_points_log : ndarray, shape (N, 2)
        Log10-space coordinates of every scatter-plot point.
    candidate_offsets : ndarray, shape (C, 2)
        Pixel-space ``(dx, dy)`` candidates generated by
        :func:`_generate_candidate_offsets`.
    """
    n_candidates = len(candidate_offsets)

    # Map pixel offsets to approximate log-space positions.
    cx = anchor_log_x + (candidate_offsets[:, 0] / _PLOT_WIDTH_PTS) * x_log_span
    cy = anchor_log_y + (candidate_offsets[:, 1] / _PLOT_HEIGHT_PTS) * y_log_span

    # -- Label repulsion (sum of inverse distances to placed labels) --------
    label_penalty = np.zeros(n_candidates)
    for lx, ly in placed_labels:
        dists = np.hypot(cx - lx, cy - ly)
        label_penalty += 1.0 / np.maximum(dists, 1e-9)

    # -- Point repulsion (sum of inverse distances to data points) ----------
    point_penalty = np.zeros(n_candidates)
    if data_points_log.size > 0:
        # Broadcast: (C,1) - (1,N) -> (C,N) pairwise distance matrix.
        dx = cx[:, np.newaxis] - data_points_log[np.newaxis, :, 0]
        dy = cy[:, np.newaxis] - data_points_log[np.newaxis, :, 1]
        dists = np.hypot(dx, dy)
        point_penalty = np.sum(1.0 / np.maximum(dists, 1e-9), axis=1)

    # -- Anchor distance (prefer positions closer to the data point) --------
    anchor_dist = np.hypot(candidate_offsets[:, 0], candidate_offsets[:, 1])

    # Weighted composite score.
    scores = 8.0 * label_penalty + 1.5 * point_penalty + 0.01 * anchor_dist

    best_idx = int(np.argmin(scores))
    return (float(candidate_offsets[best_idx, 0]),
            float(candidate_offsets[best_idx, 1]))


def annotate_outliers(
    ax,
    plot_df: pd.DataFrame,
    labeled_points_df: pd.DataFrame,
    label_color: str = "#d62728",
) -> None:
    """Annotate outlier points with customer names, minimizing label overlap.

    Labels are placed in priority order (largest absolute log-residual first)
    so that the most extreme outliers get first pick of the best positions.
    Each label's offset is chosen from a set of candidate positions arranged
    in staggered concentric rings around the anchor point.  Candidates are
    scored against both already-placed labels *and* the underlying scatter
    data to reduce visual clutter.
    """
    quantity_col = STANDARD_COLUMN_NAMES["procedure_quantity"]
    price_col = STANDARD_COLUMN_NAMES["procedure_price"]
    customer_col = STANDARD_COLUMN_NAMES["customer_name"]

    # Compute axis spans in log-space for offset-to-log conversion.
    x_vals = plot_df[quantity_col].to_numpy(dtype=float)
    y_vals = plot_df[price_col].to_numpy(dtype=float)
    x_log_span = max(np.log10(x_vals.max()) - np.log10(x_vals.min()), 1e-6)
    y_log_span = max(np.log10(y_vals.max()) - np.log10(y_vals.min()), 1e-6)

    # Pre-compute all data-point positions in log-space as an (N, 2) array
    # so the scoring function can vectorise distance calculations.
    data_points_log = np.column_stack([np.log10(x_vals), np.log10(y_vals)])

    # Generate candidate offsets once and convert to an ndarray for NumPy ops.
    candidate_offsets = np.asarray(_generate_candidate_offsets())

    # Track where labels have been placed (in log-space) so that later
    # labels can steer away from earlier ones.
    placed_labels: list[tuple[float, float]] = []

    # Sort so the most extreme outliers (highest |residual|) are placed first.
    prioritized = labeled_points_df.sort_values(
        "Log Residual",
        ascending=False,
        key=abs,
    )

    for _, row in prioritized.iterrows():
        x_val = float(row[quantity_col])
        y_val = float(row[price_col])
        anchor_lx = np.log10(x_val)
        anchor_ly = np.log10(y_val)

        offset = _select_best_offset(
            anchor_lx,
            anchor_ly,
            x_log_span,
            y_log_span,
            placed_labels,
            data_points_log,
            candidate_offsets,
        )

        # Record the chosen label position for future collision checks.
        label_lx = anchor_lx + (offset[0] / _PLOT_WIDTH_PTS) * x_log_span
        label_ly = anchor_ly + (offset[1] / _PLOT_HEIGHT_PTS) * y_log_span
        placed_labels.append((label_lx, label_ly))

        ax.annotate(
            row[customer_col],
            (x_val, y_val),
            xytext=offset,
            textcoords="offset points",
            fontsize=9,
            color=label_color,
            bbox={
                "boxstyle": "round,pad=0.2",
                "fc": "white",
                "ec": "none",
                "alpha": 0.75,
            },
            arrowprops={
                "arrowstyle": "-",
                "color": "#999999",
                "lw": 0.8,
                "alpha": 0.7,
            },
        )


# ---------------------------------------------------------------------------
# Main chart & analysis
# ---------------------------------------------------------------------------

def plot_analysis(df: pd.DataFrame, chart_title: str) -> pd.DataFrame:
    """Build the log-log price-band chart and return the full analysis table."""
    quantity_col = STANDARD_COLUMN_NAMES["procedure_quantity"]
    price_col = STANDARD_COLUMN_NAMES["procedure_price"]

    # -- Trend fitting & outlier classification ----------------------------
    _, _, trend_values, residual_std = fit_log_log_trend(df)
    analysis_df = classify_outliers(df, trend_values, residual_std)
    plot_df = analysis_df.sort_values(quantity_col).reset_index(drop=True)

    x_sorted = plot_df[quantity_col].to_numpy(dtype=float)
    trend_sorted = plot_df["Trend Price"].to_numpy(dtype=float)

    # -- Split points by band for colour-coded scatter ---------------------
    normal_df = plot_df.loc[plot_df["Band"] == "Within 1 SD"]
    lower_concern_df = plot_df.loc[plot_df["Band"] == "Between 1 and 2 SD"]
    outliers_df = plot_df.loc[plot_df["Band"] == "Outside 2 SD"]

    # Determine which points will receive customer-name labels.
    label_thresh = float(CONFIG["scatter_label_sd_threshold"])
    labeled_df = (
        plot_df.loc[plot_df["Log Residual"].abs() > label_thresh * residual_std]
        if residual_std != 0
        else plot_df.iloc[0:0]
    ).copy()

    # -- Render the chart --------------------------------------------------
    fig, ax = plt.subplots(figsize=(12, 8))

    # Scatter: normal points
    ax.scatter(
        normal_df[quantity_col], normal_df[price_col],
        alpha=0.75, s=60, label="Within 1 SD", color="#1f77b4",
    )
    # Scatter: lower-concern outliers
    if not lower_concern_df.empty:
        ax.scatter(
            lower_concern_df[quantity_col], lower_concern_df[price_col],
            alpha=0.85, s=75, label="Between 1 and 2 SD",
            color="#ff7f0e", edgecolors="black", linewidths=0.5,
        )
    # Scatter: high-concern outliers
    if not outliers_df.empty:
        ax.scatter(
            outliers_df[quantity_col], outliers_df[price_col],
            alpha=0.95, s=90, label="Outside 2 SD",
            color="#d62728", edgecolors="black",
        )

    # Trend line
    ax.plot(
        x_sorted, trend_sorted,
        color="black", linewidth=2.2, label="Log-log trend line",
    )

    # Standard-deviation band lines (+/-1 SD and +/-2 SD).
    # Only the upper line of each pair carries a legend label; the lower line
    # is drawn without one to avoid duplicate legend entries.
    bands = get_band_lines(trend_sorted, residual_std)
    if bands:
        ax.plot(x_sorted, bands["+1 SD"], "--", color="#ff7f0e", lw=1.8, label="+/- 1 SD")
        ax.plot(x_sorted, bands["-1 SD"], "--", color="#ff7f0e", lw=1.8)
        ax.plot(x_sorted, bands["+2 SD"], ":",  color="#2ca02c", lw=2.0, label="+/- 2 SD")
        ax.plot(x_sorted, bands["-2 SD"], ":",  color="#2ca02c", lw=2.0)

    ax.set_xscale("log")
    ax.set_yscale("log")

    # Annotate outlier points with customer names.
    if not labeled_df.empty:
        annotate_outliers(ax, plot_df, labeled_df)

    ax.set_title(chart_title, fontsize=15)
    ax.set_xlabel("Procedure Quantity (log scale)", fontsize=15)
    ax.set_ylabel("Procedure Price (log scale)", fontsize=15)
    ax.grid(True, which="both", linestyle=":", alpha=0.5)
    ax.legend()
    fig.tight_layout()
    plt.show()
    plt.close(fig)

    return analysis_df


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------

def save_outliers(analysis_df: pd.DataFrame, output_path: Path) -> None:
    """Save the analysis table to a formatted Excel workbook.

    Applies colour-coded severity fills, number formatting for currency /
    ratio / decimal columns, frozen header row, auto-filter, and auto-fit
    column widths.
    """
    export_df = analysis_df.sort_values(
        [
            "Severity",
            "Procedure Unit Bin",
            "Bin Outlier Methods Triggered",
            "Procedure Quantity",
            "Procedure Price",
        ],
        ascending=[True, True, False, True, False],
    )
    export_df.to_excel(
        output_path,
        index=False,
        sheet_name=CONFIG["worksheet_name"],
    )

    workbook = load_workbook(output_path)
    ws = workbook[CONFIG["worksheet_name"]]

    # -- Header styling ----------------------------------------------------
    header_fill = PatternFill(
        fill_type="solid",
        start_color=CONFIG["header_fill_color"],
        end_color=CONFIG["header_fill_color"],
    )
    header_font = Font(name="Arial", bold=True, color=CONFIG["header_font_color"])
    header_alignment = Alignment(horizontal="center", vertical="center")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # -- Severity row fills ------------------------------------------------
    severity_fills = {
        sev: PatternFill(fill_type="solid", start_color=c, end_color=c)
        for sev, c in CONFIG["severity_fill_colors"].items()
    }

    # Map header names to 1-based column indices for fast lookup.
    header_map = {
        cell.value: cell.column for cell in ws[1] if cell.value is not None
    }
    severity_col_idx = header_map.get("Severity")

    # -- Cell-level formatting (number formats and row fills) --------------
    # Note: openpyxl requires per-cell iteration; for very large exports this
    # loop dominates runtime but there is no faster built-in alternative.
    currency_cols = {
        "Procedure Price", "Trend Price", "Bin Median Price",
        "Bin IQR Lower Bound", "Bin IQR Upper Bound",
        "Bin Lower Percentile Threshold", "Bin Upper Percentile Threshold",
    }
    ratio_cols = {"Residual Ratio", "Bin Percentile Rank"}
    decimal_cols = {"Log Residual", "Bin Robust Z Score"}
    percentile_cols = {"Procedure Units Percentile", "Procedure Price Percentile"}

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        sev_value = (
            row[severity_col_idx - 1].value
            if severity_col_idx is not None
            else None
        )
        row_fill = severity_fills.get(sev_value)

        for cell in row:
            col_name = ws.cell(row=1, column=cell.column).value
            if col_name in currency_cols:
                cell.number_format = "$#,##0.00"
            elif col_name in ratio_cols:
                cell.number_format = "0.0000"
            elif col_name in decimal_cols:
                cell.number_format = "0.000"
            elif col_name in percentile_cols:
                cell.number_format = "0.0"

            if row_fill is not None:
                cell.fill = row_fill

    # -- Auto-fit column widths (capped at 40 characters) ------------------
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        max_len = max(
            len(str(cell.value)) if cell.value is not None else 0
            for cell in ws[col_letter]
        )
        ws.column_dimensions[col_letter].width = min(max_len + 2, 40)

    workbook.save(output_path)


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def parse_args() -> argparse.Namespace:
    """Parse command-line arguments."""
    parser = argparse.ArgumentParser(
        description=(
            "Generate a log-log price band analysis plot with trend, "
            "standard-deviation bands, and configurable customer labels."
        ),
    )
    parser.add_argument(
        "input_file",
        nargs="?",
        help="Input CSV or Excel file containing customer, quantity, and price columns.",
    )
    parser.add_argument(
        "--sheet",
        default=None,
        help="Excel sheet name for .xlsx or .xls inputs.",
    )
    parser.add_argument(
        "--output-outliers",
        default=None,
        help="Output Excel filename for the analysis export.",
    )
    parser.add_argument(
        "--title",
        default="Procedure Price Band Analysis",
        help="Chart title.",
    )
    return parser.parse_args()


def main() -> int:
    """Run the price band analysis workflow."""
    args = parse_args()

    try:
        input_path = Path(args.input_file) if args.input_file else choose_input_file()

        raw_df = load_input_file(input_path, args.sheet)
        cleaned_df = clean_input_data(raw_df)
        analysis_df = plot_analysis(cleaned_df, args.title)

        output_path = (
            Path(args.output_outliers)
            if args.output_outliers
            else choose_output_file(CONFIG["default_output_filename"])
        )
        save_outliers(analysis_df, output_path)

        high_concern = int((analysis_df["Band"] == "Outside 2 SD").sum())
        bin_outliers = int(analysis_df["Bin Outlier Flag"].sum())
        print("Analysis complete.  Chart displayed on screen.")
        print(f"Output file saved to: {output_path}")
        print(f"Rows analyzed: {len(analysis_df)}")
        print(f"High concern outliers (>2 SD from trend): {high_concern}")
        print(f"Customers flagged by at least one within-bin method: {bin_outliers}")
        return 0
    except Exception as exc:
        print(f"Error: {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
