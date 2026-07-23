#!/usr/bin/env python3
"""Generate a log-log price band analysis chart from customer pricing data.

The script reads a CSV or Excel file containing customer name, procedure
quantity, and procedure price columns. It standardizes the input columns,
fits a trend line in log-log space, overlays one- and two-standard-deviation
bands, displays the scatter plot on logarithmic axes, and exports the full
analysis table to a formatted Excel file.

Usage:
    python price_band_analysis.py
    python price_band_analysis.py input_file.csv
    python price_band_analysis.py input_file.xlsx --sheet "Sheet1"

Dependencies:
    pandas, numpy, matplotlib, openpyxl (for .xlsx input)
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

try:
    import tkinter as tk
    from tkinter import filedialog, messagebox
except ImportError:
    tk = None
    filedialog = None
    messagebox = None


# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------
COLUMN_ALIASES = {
    "customer_name": [
        "customer name",
        "customer group",
        "business partner",
        "customer",
        "account",
        "account name",
        "customer_name",
        "name",
    ],
    "procedure_quantity": [
        "procedure quantity",
        "procedure units",
        "quantity",
        "procedure qty",
        "qty",
        "volume",
        "procedure_quantity",
    ],
    "procedure_price": [
        "procedure price",
        "price",
        "unit price",
        "avg price",
        "procedure_price",
    ],
}

STANDARD_COLUMN_NAMES = {
    "customer_name": "Customer Name",
    "procedure_quantity": "Procedure Quantity",
    "procedure_price": "Procedure Price",
}

LABEL_OFFSETS = [
    (10, 8),
    (10, -10),
    (-10, 8),
    (-10, -10),
    (14, 0),
    (-14, 0),
    (0, 12),
    (0, -12),
    (16, 12),
    (-16, 12),
    (16, -12),
    (-16, -12),
]

CONFIG = {
    "min_procedure_units": 10,
    "max_procedure_price": 8000,
    "max_procedure_unit_bins": 4,
    "min_records_per_bin": 5,
    "band_blank_within_1sd_pricing_position": True,
    "robust_zscore_threshold": 3.5,
    "robust_iqr_multiplier": 1.5,
    "percentile_lower": 0.05,
    "percentile_upper": 0.95,
    "default_output_filename": "price_band_outliers.xlsx",
    "worksheet_name": "Price Band Analysis",
    "header_fill_color": "D9EAF7",
    "severity_fill_colors": {
        "Normal": "DDEBF7",
        "Lower Concern Outlier": "FCE4D6",
        "High Concern Outlier": "F8CBAD",
    },
}


def normalize_column_name(column_name: str) -> str:
    """Return a normalized column name for alias matching."""
    return " ".join(str(column_name).strip().lower().replace("_", " ").split())


def find_matching_column(columns: pd.Index, aliases: list[str]) -> str:
    """Return the first column whose normalized name matches an alias."""
    normalized_map = {
        normalize_column_name(column): column for column in columns
    }
    for alias in aliases:
        normalized_alias = normalize_column_name(alias)
        if normalized_alias in normalized_map:
            return normalized_map[normalized_alias]

    raise KeyError(f"Could not find a matching column for aliases: {aliases}")


def show_info_message(title: str, message: str) -> None:
    """Display an informational message using the GUI when available."""
    if messagebox and tk:
        root = tk.Tk()
        root.withdraw()
        try:
            messagebox.showinfo(title, message)
        finally:
            root.destroy()
    else:
        print(f"{title}: {message}")


def show_warning_message(title: str, message: str) -> None:
    """Display a warning message using the GUI when available."""
    if messagebox and tk:
        root = tk.Tk()
        root.withdraw()
        try:
            messagebox.showwarning(title, message)
        finally:
            root.destroy()
    else:
        print(f"{title}: {message}", file=sys.stderr)


def choose_input_file() -> Path:
    """Open a file selection dialog and return the selected input path."""
    if not tk or not filedialog:
        raise RuntimeError(
            "File dialog is not available in this Python environment. "
            "Please provide the input file as a command-line argument."
        )

    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)
    try:
        selected_file = filedialog.askopenfilename(
            title="Select input file for price band analysis",
            filetypes=[
                ("Data files", "*.csv *.xlsx *.xls"),
                ("CSV files", "*.csv"),
                ("Excel files", "*.xlsx *.xls"),
                ("All files", "*.*"),
            ],
        )
    finally:
        root.destroy()

    if not selected_file:
        raise ValueError("No input file selected.")

    return Path(selected_file)


def choose_output_file(default_name: str) -> Path:
    """Open a save dialog and return the selected Excel output path."""
    if not tk or not filedialog:
        raise RuntimeError(
            "Save dialog is not available in this Python environment. "
            "Please provide the output file as a command-line argument."
        )

    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)
    try:
        selected_file = filedialog.asksaveasfilename(
            title="Select location for Excel output",
            defaultextension=".xlsx",
            initialfile=default_name,
            filetypes=[
                ("Excel files", "*.xlsx"),
                ("All files", "*.*"),
            ],
        )
    finally:
        root.destroy()

    if not selected_file:
        raise ValueError("No output file selected.")

    return Path(selected_file)


def load_excel_sheet(
    file_path: Path,
    sheet_name: str | None = None,
) -> pd.DataFrame:
    """Load the requested Excel sheet or infer the most likely data sheet."""
    if sheet_name:
        return pd.read_excel(file_path, sheet_name=sheet_name)

    workbook = pd.read_excel(file_path, sheet_name=None)
    normalized_quantity_aliases = {
        normalize_column_name(alias)
        for alias in COLUMN_ALIASES["procedure_quantity"]
    }

    for candidate_df in workbook.values():
        normalized_columns = {
            normalize_column_name(column) for column in candidate_df.columns
        }
        if normalized_quantity_aliases & normalized_columns:
            return candidate_df

    first_sheet_name = next(iter(workbook))
    return workbook[first_sheet_name]


def load_input_file(
    file_path: Path,
    sheet_name: str | None = None,
) -> pd.DataFrame:
    """Load a CSV or Excel input file into a DataFrame."""
    suffix = file_path.suffix.lower()
    if suffix == ".csv":
        return pd.read_csv(file_path)
    if suffix in {".xlsx", ".xls"}:
        return load_excel_sheet(file_path, sheet_name)

    raise ValueError("Unsupported file type. Use CSV or Excel (.xlsx/.xls).")


def standardize_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Map source column names to the standard analysis column names."""
    mapped_columns = {
        key: find_matching_column(df.columns, aliases)
        for key, aliases in COLUMN_ALIASES.items()
    }

    standardized_df = df[
        [
            mapped_columns["customer_name"],
            mapped_columns["procedure_quantity"],
            mapped_columns["procedure_price"],
        ]
    ].copy()
    standardized_df.columns = [
        STANDARD_COLUMN_NAMES["customer_name"],
        STANDARD_COLUMN_NAMES["procedure_quantity"],
        STANDARD_COLUMN_NAMES["procedure_price"],
    ]
    return standardized_df


def clean_input_data(df: pd.DataFrame) -> pd.DataFrame:
    """Standardize, validate, and filter the input data for analysis."""
    cleaned_df = standardize_columns(df)

    customer_col = STANDARD_COLUMN_NAMES["customer_name"]
    quantity_col = STANDARD_COLUMN_NAMES["procedure_quantity"]
    price_col = STANDARD_COLUMN_NAMES["procedure_price"]

    cleaned_df[customer_col] = cleaned_df[customer_col].astype(str).str.strip()
    cleaned_df.loc[
        cleaned_df[customer_col].isin(["", "nan", "None"]),
        customer_col,
    ] = "Unknown"

    cleaned_df[quantity_col] = pd.to_numeric(cleaned_df[quantity_col], errors="coerce")
    cleaned_df[price_col] = pd.to_numeric(cleaned_df[price_col], errors="coerce")

    invalid_numeric_mask = (
        cleaned_df[quantity_col].isna() | cleaned_df[price_col].isna()
    )
    dropped_count = int(invalid_numeric_mask.sum())

    cleaned_df = cleaned_df.loc[~invalid_numeric_mask].copy()
    cleaned_df = cleaned_df.loc[
        (cleaned_df[quantity_col] >= CONFIG["min_procedure_units"])
        & (cleaned_df[price_col] > 0)
    ].copy()
    cleaned_df = cleaned_df.loc[
        cleaned_df[price_col] <= CONFIG["max_procedure_price"]
    ].copy()

    if cleaned_df.empty:
        message = (
            "No valid data remains after cleaning and filtering. "
            f"Records must have Procedure Quantity >= {CONFIG['min_procedure_units']} "
            f"and Procedure Price <= {CONFIG['max_procedure_price']}."
        )
        if dropped_count > 0:
            message += (
                f"\n\nRows dropped due to invalid numeric values: {dropped_count}"
            )
        show_warning_message("No Valid Data", message)
        raise ValueError("No valid rows found after cleaning input data.")

    if dropped_count > 0:
        show_info_message(
            "Data Cleaning Complete",
            "Loaded successfully.\n"
            f"Rows dropped due to invalid numeric values: {dropped_count}",
        )

    return cleaned_df


def fit_log_log_trend(
    df: pd.DataFrame,
) -> tuple[np.ndarray, np.ndarray, np.ndarray, float]:
    """Fit a linear regression in log-log space and return trend outputs."""
    quantity_col = STANDARD_COLUMN_NAMES["procedure_quantity"]
    price_col = STANDARD_COLUMN_NAMES["procedure_price"]

    x_values = df[quantity_col].to_numpy(dtype=float)
    y_values = df[price_col].to_numpy(dtype=float)

    if len(df) < 2:
        raise ValueError("At least two rows are required to fit a trend line.")

    log_x = np.log10(x_values)
    log_y = np.log10(y_values)
    slope, intercept = np.polyfit(log_x, log_y, 1)
    trend_log = slope * log_x + intercept
    residual_std = float(np.std(log_y - trend_log, ddof=1)) if len(df) > 2 else 0.0
    trend_values = np.power(10.0, trend_log)

    return x_values, y_values, trend_values, residual_std


def build_procedure_unit_bins(
    quantity_series: pd.Series,
) -> tuple[pd.Series, pd.DataFrame]:
    """Assign each row to a procedure unit bin and return bin range metadata."""
    bin_count = min(CONFIG["max_procedure_unit_bins"], max(len(quantity_series), 1))

    if bin_count <= 1:
        bin_labels = pd.Series(["Bin 1"] * len(quantity_series), index=quantity_series.index)
    else:
        rank_values = quantity_series.rank(method="first")
        raw_bins = pd.qcut(rank_values, q=bin_count, labels=False, duplicates="drop")
        if raw_bins is None:
            bin_labels = pd.Series(
                ["Bin 1"] * len(quantity_series),
                index=quantity_series.index,
            )
        else:
            bin_labels = raw_bins.astype(int).add(1).map(
                lambda value: f"Bin {value}"
            )

    bin_summary = (
        pd.DataFrame(
            {
                "Procedure Unit Bin": bin_labels,
                "Procedure Quantity": quantity_series,
            }
        )
        .groupby("Procedure Unit Bin", sort=True)["Procedure Quantity"]
        .agg(["min", "max"])
        .reset_index()
    )
    bin_summary["Procedure Unit Bin Range"] = bin_summary.apply(
        lambda row: f"{row['min']:,.0f} to {row['max']:,.0f}",
        axis=1,
    )

    return bin_labels, bin_summary[["Procedure Unit Bin", "Procedure Unit Bin Range"]]


def calculate_robust_z_scores(series: pd.Series) -> pd.Series:
    """Return robust z-scores based on the median absolute deviation."""
    median_value = series.median()
    mad = np.median(np.abs(series - median_value))

    if mad == 0 or np.isnan(mad):
        return pd.Series(0.0, index=series.index)

    return 0.6745 * (series - median_value) / mad


def append_bin_outlier_metrics(analysis_df: pd.DataFrame) -> pd.DataFrame:
    """Add procedure-unit bin labels and within-bin outlier diagnostics."""
    price_col = STANDARD_COLUMN_NAMES["procedure_price"]
    quantity_col = STANDARD_COLUMN_NAMES["procedure_quantity"]

    enriched_df = analysis_df.copy()
    bin_labels, bin_summary = build_procedure_unit_bins(enriched_df[quantity_col])
    enriched_df["Procedure Unit Bin"] = bin_labels
    enriched_df = enriched_df.merge(bin_summary, on="Procedure Unit Bin", how="left")

    enriched_df["Bin Peer Count"] = 0
    enriched_df["Bin Median Price"] = np.nan
    enriched_df["Bin Robust Z Score"] = 0.0
    enriched_df["Bin Robust Z Score Outlier"] = False
    enriched_df["Bin IQR Lower Bound"] = np.nan
    enriched_df["Bin IQR Upper Bound"] = np.nan
    enriched_df["Bin Robust IQR Outlier"] = False
    enriched_df["Bin Percentile Rank"] = np.nan
    enriched_df["Bin Lower Percentile Threshold"] = np.nan
    enriched_df["Bin Upper Percentile Threshold"] = np.nan
    enriched_df["Bin Percentile Extreme"] = False
    enriched_df["Bin Outlier Methods Triggered"] = 0
    enriched_df["Bin Outlier Flag"] = False
    enriched_df["Bin Outlier Methods"] = ""

    for _, bin_index in enriched_df.groupby("Procedure Unit Bin").groups.items():
        bin_prices = enriched_df.loc[bin_index, price_col]
        peer_count = len(bin_prices)
        q1 = bin_prices.quantile(0.25)
        q3 = bin_prices.quantile(0.75)
        iqr = q3 - q1
        lower_bound = q1 - (CONFIG["robust_iqr_multiplier"] * iqr)
        upper_bound = q3 + (CONFIG["robust_iqr_multiplier"] * iqr)
        robust_z_scores = calculate_robust_z_scores(bin_prices)
        percentile_ranks = bin_prices.rank(method="average", pct=True)
        lower_threshold = bin_prices.quantile(CONFIG["percentile_lower"])
        upper_threshold = bin_prices.quantile(CONFIG["percentile_upper"])

        robust_z_mask = robust_z_scores.abs() > CONFIG["robust_zscore_threshold"]
        iqr_mask = (bin_prices < lower_bound) | (bin_prices > upper_bound)
        percentile_mask = (
            (percentile_ranks <= CONFIG["percentile_lower"])
            | (percentile_ranks >= CONFIG["percentile_upper"])
        )

        if peer_count < CONFIG["min_records_per_bin"]:
            robust_z_mask = pd.Series(False, index=bin_prices.index)
            iqr_mask = pd.Series(False, index=bin_prices.index)
            percentile_mask = pd.Series(False, index=bin_prices.index)

        methods_triggered = (
            robust_z_mask.astype(int)
            + iqr_mask.astype(int)
            + percentile_mask.astype(int)
        )
        method_labels = pd.DataFrame(
            {
                "robust_z": robust_z_mask,
                "robust_iqr": iqr_mask,
                "percentile_extreme": percentile_mask,
            }
        )

        enriched_df.loc[bin_index, "Bin Peer Count"] = peer_count
        enriched_df.loc[bin_index, "Bin Median Price"] = bin_prices.median()
        enriched_df.loc[bin_index, "Bin Robust Z Score"] = robust_z_scores
        enriched_df.loc[bin_index, "Bin Robust Z Score Outlier"] = robust_z_mask
        enriched_df.loc[bin_index, "Bin IQR Lower Bound"] = lower_bound
        enriched_df.loc[bin_index, "Bin IQR Upper Bound"] = upper_bound
        enriched_df.loc[bin_index, "Bin Robust IQR Outlier"] = iqr_mask
        enriched_df.loc[bin_index, "Bin Percentile Rank"] = percentile_ranks
        enriched_df.loc[bin_index, "Bin Lower Percentile Threshold"] = lower_threshold
        enriched_df.loc[bin_index, "Bin Upper Percentile Threshold"] = upper_threshold
        enriched_df.loc[bin_index, "Bin Percentile Extreme"] = percentile_mask
        enriched_df.loc[bin_index, "Bin Outlier Methods Triggered"] = methods_triggered
        enriched_df.loc[bin_index, "Bin Outlier Flag"] = methods_triggered > 0
        enriched_df.loc[bin_index, "Bin Outlier Methods"] = method_labels.apply(
            lambda row: ", ".join(
                label
                for label, triggered in [
                    ("Robust Z score", row["robust_z"]),
                    ("Robust IQR", row["robust_iqr"]),
                    ("Percentile extremes", row["percentile_extreme"]),
                ]
                if triggered
            ),
            axis=1,
        )

    return enriched_df


def classify_outliers(
    df: pd.DataFrame,
    trend_values: np.ndarray,
    residual_std: float,
) -> pd.DataFrame:
    """Classify records by deviation from the log-log trend."""
    price_col = STANDARD_COLUMN_NAMES["procedure_price"]

    analysis_df = df.copy()
    analysis_df["Trend Price"] = trend_values
    analysis_df["Log Residual"] = (
        np.log10(analysis_df[price_col]) - np.log10(analysis_df["Trend Price"])
    )
    analysis_df["Residual Ratio"] = (
        analysis_df[price_col] / analysis_df["Trend Price"]
    )
    analysis_df["Band"] = "Within 1 SD"
    analysis_df["Severity"] = "Normal"
    analysis_df["Pricing Position"] = ""

    if residual_std != 0:
        abs_log_residual = analysis_df["Log Residual"].abs()
        one_sd_mask = (
            (abs_log_residual > residual_std)
            & (abs_log_residual <= 2 * residual_std)
        )
        two_sd_mask = abs_log_residual > 2 * residual_std

        analysis_df.loc[one_sd_mask, "Band"] = "Between 1 and 2 SD"
        analysis_df.loc[two_sd_mask, "Band"] = "Outside 2 SD"
        analysis_df.loc[one_sd_mask, "Severity"] = "Lower Concern Outlier"
        analysis_df.loc[two_sd_mask, "Severity"] = "High Concern Outlier"

        outside_one_sd_mask = abs_log_residual > residual_std
        analysis_df.loc[
            outside_one_sd_mask & (analysis_df[price_col] < analysis_df["Trend Price"]),
            "Pricing Position",
        ] = "Potentially Underpriced"
        analysis_df.loc[
            outside_one_sd_mask & (analysis_df[price_col] > analysis_df["Trend Price"]),
            "Pricing Position",
        ] = "Potentially Overpriced"

    if not CONFIG["band_blank_within_1sd_pricing_position"]:
        analysis_df.loc[
            analysis_df["Band"] == "Within 1 SD", "Pricing Position"
        ] = "In Line"

    return append_bin_outlier_metrics(analysis_df)


def get_band_lines(
    trend_values: np.ndarray,
    residual_std: float,
) -> dict[str, np.ndarray]:
    """Return multiplicative one- and two-standard-deviation band lines."""
    if residual_std == 0:
        return {}

    sd_factor_1 = np.power(10.0, residual_std)
    sd_factor_2 = np.power(10.0, 2 * residual_std)
    return {
        "+1 SD": trend_values * sd_factor_1,
        "-1 SD": trend_values / sd_factor_1,
        "+2 SD": trend_values * sd_factor_2,
        "-2 SD": trend_values / sd_factor_2,
    }


def choose_label_offset(
    x_value: float,
    y_value: float,
    x_log_span: float,
    y_log_span: float,
    placed_positions: list[tuple[float, float]],
) -> tuple[int, int]:
    """Choose a label offset that minimizes overlap with existing labels."""
    x_log = np.log10(x_value)
    y_log = np.log10(y_value)
    best_offset = LABEL_OFFSETS[0]
    best_score = None

    for dx, dy in LABEL_OFFSETS:
        candidate_x = x_log + (dx / 220.0) * x_log_span
        candidate_y = y_log + (dy / 220.0) * y_log_span
        score = 0.0

        for placed_x, placed_y in placed_positions:
            distance = np.hypot(candidate_x - placed_x, candidate_y - placed_y)
            score += 1.0 / max(distance, 1e-6)

        if best_score is None or score < best_score:
            best_score = score
            best_offset = (dx, dy)

    return best_offset


def annotate_outliers(ax, plot_df: pd.DataFrame, outliers_df: pd.DataFrame) -> None:
    """Annotate outlier points while reducing label overlap where possible."""
    quantity_col = STANDARD_COLUMN_NAMES["procedure_quantity"]
    price_col = STANDARD_COLUMN_NAMES["procedure_price"]
    customer_col = STANDARD_COLUMN_NAMES["customer_name"]

    x_min = float(plot_df[quantity_col].min())
    x_max = float(plot_df[quantity_col].max())
    y_min = float(plot_df[price_col].min())
    y_max = float(plot_df[price_col].max())

    x_log_span = max(np.log10(x_max) - np.log10(x_min), 1e-6)
    y_log_span = max(np.log10(y_max) - np.log10(y_min), 1e-6)
    placed_positions: list[tuple[float, float]] = []

    prioritized_outliers = outliers_df.copy()
    prioritized_outliers["Label Priority"] = prioritized_outliers["Log Residual"].abs()

    for _, row in prioritized_outliers.sort_values(
        "Label Priority",
        ascending=False,
    ).iterrows():
        x_value = float(row[quantity_col])
        y_value = float(row[price_col])
        offset = choose_label_offset(
            x_value,
            y_value,
            x_log_span,
            y_log_span,
            placed_positions,
        )

        label_x = np.log10(x_value) + (offset[0] / 220.0) * x_log_span
        label_y = np.log10(y_value) + (offset[1] / 220.0) * y_log_span
        placed_positions.append((label_x, label_y))

        ax.annotate(
            row[customer_col],
            (x_value, y_value),
            xytext=offset,
            textcoords="offset points",
            fontsize=9,
            color="#d62728",
            bbox={
                "boxstyle": "round,pad=0.2",
                "fc": "white",
                "ec": "none",
                "alpha": 0.75,
            },
            arrowprops={
                "arrowstyle": "-",
                "color": "#999999",
                "lw": 0.8,
                "alpha": 0.7,
            },
        )


def plot_analysis(df: pd.DataFrame, chart_title: str) -> pd.DataFrame:
    """Display the log-log price band chart and return the analysis table."""
    quantity_col = STANDARD_COLUMN_NAMES["procedure_quantity"]
    price_col = STANDARD_COLUMN_NAMES["procedure_price"]

    _, _, trend_values, residual_std = fit_log_log_trend(df)
    analysis_df = classify_outliers(df, trend_values, residual_std)
    plot_df = analysis_df.sort_values(quantity_col).reset_index(drop=True)

    x_sorted = plot_df[quantity_col].to_numpy(dtype=float)
    trend_sorted = plot_df["Trend Price"].to_numpy(dtype=float)
    normal_df = plot_df.loc[plot_df["Band"] == "Within 1 SD"]
    lower_concern_df = plot_df.loc[
        plot_df["Band"] == "Between 1 and 2 SD"
    ].copy()
    outliers_df = plot_df.loc[plot_df["Band"] == "Outside 2 SD"].copy()

    fig, ax = plt.subplots(figsize=(12, 8))
    ax.scatter(
        normal_df[quantity_col],
        normal_df[price_col],
        alpha=0.75,
        s=60,
        label="Within 1 SD",
        color="#1f77b4",
    )

    if not lower_concern_df.empty:
        ax.scatter(
            lower_concern_df[quantity_col],
            lower_concern_df[price_col],
            alpha=0.85,
            s=75,
            label="Between 1 and 2 SD",
            color="#ff7f0e",
            edgecolors="black",
            linewidths=0.5,
        )

    if not outliers_df.empty:
        ax.scatter(
            outliers_df[quantity_col],
            outliers_df[price_col],
            alpha=0.95,
            s=90,
            label="Outside 2 SD",
            color="#d62728",
            edgecolors="black",
        )

    ax.plot(
        x_sorted,
        trend_sorted,
        color="black",
        linewidth=2.2,
        label="Log-log trend line",
    )

    band_lines = get_band_lines(trend_sorted, residual_std)
    if band_lines:
        ax.plot(
            x_sorted,
            band_lines["+1 SD"],
            linestyle="--",
            color="#ff7f0e",
            linewidth=1.8,
            label="+1 SD",
        )
        ax.plot(
            x_sorted,
            band_lines["-1 SD"],
            linestyle="--",
            color="#ff7f0e",
            linewidth=1.8,
            label="-1 SD",
        )
        ax.plot(
            x_sorted,
            band_lines["+2 SD"],
            linestyle=":",
            color="#2ca02c",
            linewidth=2.0,
            label="+2 SD",
        )
        ax.plot(
            x_sorted,
            band_lines["-2 SD"],
            linestyle=":",
            color="#2ca02c",
            linewidth=2.0,
            label="-2 SD",
        )

    ax.set_xscale("log")
    ax.set_yscale("log")

    if not outliers_df.empty:
        annotate_outliers(ax, plot_df, outliers_df)

    ax.set_title(chart_title)
    ax.set_xlabel("Procedure Quantity (log scale)")
    ax.set_ylabel("Procedure Price (log scale)")
    ax.grid(True, which="both", linestyle=":", alpha=0.5)
    ax.legend()
    fig.tight_layout()
    plt.show()
    plt.close(fig)

    return analysis_df


def save_outliers(analysis_df: pd.DataFrame, output_outliers: Path) -> None:
    """Save the analysis table to a formatted Excel workbook."""
    export_df = analysis_df.copy()
    export_df = export_df.sort_values(
        [
            "Severity",
            "Procedure Unit Bin",
            "Bin Outlier Methods Triggered",
            "Procedure Quantity",
            "Procedure Price",
        ],
        ascending=[True, True, False, True, False],
    )
    export_df.to_excel(
        output_outliers,
        index=False,
        sheet_name=CONFIG["worksheet_name"],
    )

    workbook = load_workbook(output_outliers)
    worksheet = workbook[CONFIG["worksheet_name"]]

    header_fill = PatternFill(
        fill_type="solid",
        start_color=CONFIG["header_fill_color"],
        end_color=CONFIG["header_fill_color"],
    )
    header_font = Font(name="Arial", bold=True, color="000000")
    header_alignment = Alignment(horizontal="center", vertical="center")

    severity_fills = {
        severity: PatternFill(
            fill_type="solid",
            start_color=color,
            end_color=color,
        )
        for severity, color in CONFIG["severity_fill_colors"].items()
    }

    header_map = {
        cell.value: cell.column for cell in worksheet[1] if cell.value is not None
    }

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    currency_columns = {
        "Procedure Price",
        "Trend Price",
        "Bin Median Price",
        "Bin IQR Lower Bound",
        "Bin IQR Upper Bound",
        "Bin Lower Percentile Threshold",
        "Bin Upper Percentile Threshold",
    }
    ratio_columns = {"Residual Ratio", "Bin Percentile Rank"}
    decimal_columns = {"Log Residual", "Bin Robust Z Score"}
    severity_column_index = header_map.get("Severity")

    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
        severity_value = None
        if severity_column_index is not None:
            severity_value = row[severity_column_index - 1].value

        row_fill = severity_fills.get(severity_value)

        for cell in row:
            column_name = worksheet.cell(row=1, column=cell.column).value
            if column_name in currency_columns:
                cell.number_format = "$#,##0.00"
            elif column_name in ratio_columns:
                cell.number_format = "0.0000"
            elif column_name in decimal_columns:
                cell.number_format = "0.000"

            if row_fill is not None:
                cell.fill = row_fill

    for column_index in range(1, worksheet.max_column + 1):
        column_letter = get_column_letter(column_index)
        max_length = max(
            len(str(cell.value)) if cell.value is not None else 0
            for cell in worksheet[column_letter]
        )
        worksheet.column_dimensions[column_letter].width = min(max_length + 2, 40)

    workbook.save(output_outliers)


def parse_args() -> argparse.Namespace:
    """Parse command-line arguments for the analysis script."""
    parser = argparse.ArgumentParser(
        description=(
            "Generate a log-log price band analysis plot with trend and "
            "standard-deviation bands."
        )
    )
    parser.add_argument(
        "input_file",
        nargs="?",
        help="Input CSV or Excel file containing customer, quantity, and price columns.",
    )
    parser.add_argument(
        "--sheet",
        default=None,
        help="Excel sheet name for .xlsx or .xls inputs.",
    )
    parser.add_argument(
        "--output-outliers",
        default=None,
        help="Output Excel filename for the analysis export.",
    )
    parser.add_argument(
        "--title",
        default="Procedure Price Band Analysis",
        help="Chart title.",
    )
    return parser.parse_args()


def main() -> int:
    """Run the price band analysis workflow."""
    args = parse_args()

    try:
        input_path = Path(args.input_file) if args.input_file else choose_input_file()

        raw_df = load_input_file(input_path, args.sheet)
        cleaned_df = clean_input_data(raw_df)
        analysis_df = plot_analysis(cleaned_df, args.title)

        output_path = (
            Path(args.output_outliers)
            if args.output_outliers
            else choose_output_file(CONFIG["default_output_filename"])
        )
        save_outliers(analysis_df, output_path)

        high_concern_count = int(
            (analysis_df["Band"] == "Outside 2 SD").sum()
        )
        bin_outlier_count = int(analysis_df["Bin Outlier Flag"].sum())
        print("Analysis complete. Chart displayed on screen.")
        print(f"Output file saved to: {output_path}")
        print(f"Rows analyzed: {len(analysis_df)}")
        print(
            "High concern outliers identified (>2 SD from trend): "
            f"{high_concern_count}"
        )
        print(
            "Customers flagged by at least one within-bin method: "
            f"{bin_outlier_count}"
        )
        return 0
    except Exception as exc:
        print(f"Error: {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
